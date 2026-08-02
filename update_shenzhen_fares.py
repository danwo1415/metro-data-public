#!/usr/bin/env python3
"""Download and parse Shenzhen's official ordinary-class metro fare matrix.

The source is the Shenzhen Development and Reform Commission's 2026 fare-table
workbook. The script discovers the attachment from official notice pages, maps
station names to the stable codes in shenzhen-data.json, validates coverage,
and updates shenzhen-data.json plus shenzhen-version.json atomically.
"""
from __future__ import annotations

import hashlib
import io
import json
import os
import re
import sys
import unicodedata
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

DATA_FILE = Path("shenzhen-data.json")
VERSION_FILE = Path("shenzhen-version.json")
REPORT_FILE = Path("shenzhen-fares-report.json")
TZ = timezone(timedelta(hours=8))
HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; MetroRoutePlanner/1.0; +https://github.com/danwo1415/metro-data-public)",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.7",
}
NOTICE_PAGES = [
    "https://fgw.sz.gov.cn/gkmlpt/content/12/12853/post_12853584.html",
    "https://www.sz.gov.cn/szzt2010/zdlyzl/sfxx/bz/fw/content/post_12853619.html",
]
# Official fallback attachment URLs observed on the two government notice pages.
FALLBACK_XLSX = [
    "https://fgw.sz.gov.cn/attachment/1/1723/1723434/12853584.xlsx",
    "https://www.sz.gov.cn/attachment/1/1723/1723434/12853619.xlsx",
    "https://fgw.sz.gov.cn/attachment/1/1723/1723434/12849099.xlsx",
]


def norm(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).lower().strip()
    text = re.sub(r"[（(].*?[）)]", "", text)
    text = text.replace("地铁", "").replace("地鐵", "").replace("metro", "")
    text = re.sub(r"station$", "", text)
    text = re.sub(r"站$", "", text)
    return re.sub(r"[^a-z0-9\u3400-\u9fff]", "", text)


def number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        v = float(value)
    else:
        m = re.search(r"\d+(?:\.\d+)?", str(value).replace(",", ""))
        if not m:
            return None
        v = float(m.group())
    return v if 0 < v < 100 else None


def build_lookup(data: dict[str, Any]) -> tuple[dict[str, str], dict[str, list[str]]]:
    lookup: dict[str, str] = {}
    collisions: dict[str, list[str]] = {}
    all_names: dict[str, set[str]] = {}
    for line in data["L"].values():
        for code, trad in line["stations"]:
            names = {
                trad,
                data.get("simplified", {}).get(code, ""),
                data.get("english", {}).get(code, ""),
            }
            for name in names:
                key = norm(name)
                if key:
                    all_names.setdefault(key, set()).add(code)
    for key, codes in all_names.items():
        if len(codes) == 1:
            lookup[key] = next(iter(codes))
        else:
            collisions[key] = sorted(codes)
    return lookup, collisions


def map_station(value: Any, lookup: dict[str, str]) -> str | None:
    key = norm(value)
    if not key:
        return None
    if key in lookup:
        return lookup[key]
    # Official workbooks occasionally append a line number or "枢纽" marker.
    variants = [
        re.sub(r"(?:\d+号线|\d+號綫|线|綫|枢纽|樞紐)$", "", key),
        key.replace("深圳北", "深圳北站"),
    ]
    for item in variants:
        if item in lookup:
            return lookup[item]
    # Use a unique containment match only for long names to avoid false matches.
    if len(key) >= 4:
        hits = {code for name, code in lookup.items() if key in name or name in key}
        if len(hits) == 1:
            return next(iter(hits))
    return None


def discover_xlsx(session: requests.Session) -> list[str]:
    urls: list[str] = []
    for page in NOTICE_PAGES:
        try:
            r = session.get(page, headers=HEADERS, timeout=35)
            r.raise_for_status()
            soup = BeautifulSoup(r.text, "lxml")
            for a in soup.find_all("a", href=True):
                text = a.get_text(" ", strip=True)
                href = urljoin(page, a["href"])
                if ".xlsx" in href.lower() and ("普通" in text or "普通" in href or not urls):
                    urls.append(href)
        except Exception:
            continue
    urls.extend(FALLBACK_XLSX)
    return list(dict.fromkeys(urls))


def download_workbook(session: requests.Session) -> tuple[bytes, str]:
    errors: list[str] = []
    for url in discover_xlsx(session):
        try:
            r = session.get(url, headers=HEADERS, timeout=60)
            r.raise_for_status()
            content = r.content
            if len(content) < 5000 or not content.startswith(b"PK"):
                raise RuntimeError(f"not an XLSX payload ({len(content)} bytes)")
            return content, url
        except Exception as exc:
            errors.append(f"{url}: {exc}")
    raise RuntimeError("Unable to download official fare workbook: " + " | ".join(errors[-5:]))


def parse_long_form(ws: Any, lookup: dict[str, str]) -> tuple[dict[str, dict[str, float]], dict[str, Any]]:
    fares: dict[str, dict[str, float]] = {}
    best = {"pairs": 0, "headerRow": None}
    max_r = min(ws.max_row, 2000)
    max_c = min(ws.max_column, 30)
    for r in range(1, min(max_r, 80) + 1):
        headers = [norm(ws.cell(r, c).value) for c in range(1, max_c + 1)]
        origin_col = next((i + 1 for i, h in enumerate(headers) if h in {"起点", "起點", "出发站", "出發站", "origin"}), None)
        dest_col = next((i + 1 for i, h in enumerate(headers) if h in {"终点", "終點", "到达站", "到達站", "destination"}), None)
        fare_col = next((i + 1 for i, h in enumerate(headers) if "票价" in h or "票價" in h or h == "fare"), None)
        if not (origin_col and dest_col and fare_col):
            continue
        temp: dict[str, dict[str, float]] = {}
        pairs = 0
        for rr in range(r + 1, max_r + 1):
            a = map_station(ws.cell(rr, origin_col).value, lookup)
            b = map_station(ws.cell(rr, dest_col).value, lookup)
            v = number(ws.cell(rr, fare_col).value)
            if not (a and b and v is not None) or a == b:
                continue
            x, y = sorted((a, b))
            temp.setdefault(x, {})[y] = int(v) if v.is_integer() else round(v, 1)
            pairs += 1
        if pairs > best["pairs"]:
            fares, best = temp, {"pairs": pairs, "headerRow": r}
    return fares, best


def parse_matrix(ws: Any, lookup: dict[str, str]) -> tuple[dict[str, dict[str, float]], dict[str, Any]]:
    max_r, max_c = min(ws.max_row, 700), min(ws.max_column, 700)
    row_candidates: list[tuple[int, dict[int, str]]] = []
    for r in range(1, min(max_r, 60) + 1):
        mapped = {c: code for c in range(1, max_c + 1) if (code := map_station(ws.cell(r, c).value, lookup))}
        if len(mapped) >= 20:
            row_candidates.append((r, mapped))
    col_candidates: list[tuple[int, dict[int, str]]] = []
    for c in range(1, min(max_c, 60) + 1):
        mapped = {r: code for r in range(1, max_r + 1) if (code := map_station(ws.cell(r, c).value, lookup))}
        if len(mapped) >= 20:
            col_candidates.append((c, mapped))
    row_candidates.sort(key=lambda x: len(x[1]), reverse=True)
    col_candidates.sort(key=lambda x: len(x[1]), reverse=True)

    best_fares: dict[str, dict[str, float]] = {}
    best_info: dict[str, Any] = {"pairs": 0}
    for header_row, columns in row_candidates[:8]:
        for header_col, rows in col_candidates[:8]:
            temp: dict[str, dict[str, float]] = {}
            pairs = 0
            mapped_codes: set[str] = set()
            for rr, a in rows.items():
                if rr <= header_row:
                    continue
                for cc, b in columns.items():
                    if cc <= header_col or a == b:
                        continue
                    v = number(ws.cell(rr, cc).value)
                    if v is None:
                        continue
                    x, y = sorted((a, b))
                    temp.setdefault(x, {})[y] = int(v) if v.is_integer() else round(v, 1)
                    pairs += 1
                    mapped_codes.update((a, b))
            unique_pairs = sum(len(v) for v in temp.values())
            if unique_pairs > best_info.get("pairs", 0):
                best_fares = temp
                best_info = {
                    "pairs": unique_pairs,
                    "rawNumericCells": pairs,
                    "stations": len(mapped_codes),
                    "headerRow": header_row,
                    "headerColumn": header_col,
                    "columnHeaders": len(columns),
                    "rowHeaders": len(rows),
                }
    return best_fares, best_info


def parse_workbook(content: bytes, lookup: dict[str, str]) -> tuple[dict[str, dict[str, float]], dict[str, Any]]:
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=False)
    best_fares: dict[str, dict[str, float]] = {}
    report: dict[str, Any] = {"sheets": []}
    for ws in wb.worksheets:
        matrix, matrix_info = parse_matrix(ws, lookup)
        long_form, long_info = parse_long_form(ws, lookup)
        chosen, info, mode = (matrix, matrix_info, "matrix") if matrix_info.get("pairs", 0) >= long_info.get("pairs", 0) else (long_form, long_info, "long-form")
        sheet_report = {"sheet": ws.title, "rows": ws.max_row, "columns": ws.max_column, "mode": mode, **info}
        report["sheets"].append(sheet_report)
        if sum(len(v) for v in chosen.values()) > sum(len(v) for v in best_fares.values()):
            best_fares = chosen
            report["selectedSheet"] = ws.title
            report["selectedMode"] = mode
            report["selectedInfo"] = info
    return best_fares, report


def canonical(fares: dict[str, dict[str, float]]) -> str:
    return json.dumps(fares, sort_keys=True, separators=(",", ":"), ensure_ascii=False)


def main() -> int:
    if not DATA_FILE.exists() or not VERSION_FILE.exists():
        print("ERROR: shenzhen-data.json or shenzhen-version.json is missing", file=sys.stderr)
        return 2
    data = json.loads(DATA_FILE.read_text(encoding="utf-8"))
    manifest = json.loads(VERSION_FILE.read_text(encoding="utf-8"))
    lookup, collisions = build_lookup(data)
    session = requests.Session()

    print("[1/4] Downloading official Shenzhen ordinary-class fare workbook...", flush=True)
    try:
        content, source_url = download_workbook(session)
    except Exception as exc:
        REPORT_FILE.write_text(json.dumps({"error": str(exc)}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        print(f"ERROR: {exc}", file=sys.stderr)
        return 3

    print("[2/4] Parsing official fare matrix...", flush=True)
    fares, parse_report = parse_workbook(content, lookup)
    pair_count = sum(len(v) for v in fares.values())
    station_codes = {code for origin, row in fares.items() for destination in row for code in (origin, destination)}
    report = {
        "sourceUrl": source_url,
        "downloadBytes": len(content),
        "pairs": pair_count,
        "stations": len(station_codes),
        "lookupCollisions": collisions,
        **parse_report,
    }
    REPORT_FILE.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(f"[3/4] Validating: stations={len(station_codes)} pairs={pair_count}", flush=True)
    if len(station_codes) < 150 or pair_count < 10000:
        print("ERROR: fare-table coverage is below the safety threshold; data files not replaced. See shenzhen-fares-report.json", file=sys.stderr)
        return 4

    old = data.get("fares") or {}
    if canonical(old) == canonical(fares):
        print("[4/4] SUCCESS: official fares unchanged", flush=True)
        return 0

    now = datetime.now(TZ)
    digest = hashlib.sha256(canonical(fares).encode("utf-8")).hexdigest()[:10]
    version = f"2.5.1-sz-data-{now:%Y%m%d%H%M}-{digest}"
    data["fares"] = fares
    data["fareMetadata"] = {
        "type": "ordinary-class",
        "currency": "CNY",
        "source": "Shenzhen Development and Reform Commission official network fare workbook",
        "sourceUrl": source_url,
        "updatedAt": now.isoformat(),
        "stations": len(station_codes),
        "pairs": pair_count,
    }
    data["version"] = version
    data["updatedAt"] = now.date().isoformat()
    manifest["version"] = version
    manifest["updatedAt"] = now.date().isoformat()

    tmp_data = DATA_FILE.with_suffix(".json.tmp")
    tmp_version = VERSION_FILE.with_suffix(".json.tmp")
    tmp_data.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    tmp_version.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    os.replace(tmp_data, DATA_FILE)
    os.replace(tmp_version, VERSION_FILE)
    print(f"[4/4] SUCCESS: version={version} stations={len(station_codes)} pairs={pair_count}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
